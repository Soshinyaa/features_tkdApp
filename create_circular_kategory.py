import os
import re
import openpyxl
from openpyxl import Workbook
file = 'protocol.xlsx'


def create_circular_kategory(file):
    wb = openpyxl.load_workbook(file)
    for ws in wb:
        values = [
            '№', 'Ф.И.О.', 'Дата рождения', 'Спортивная квалификация', 'Техническая квалификация',
            'Область, край, республика', 'Федеральный округ', 'ДСО, ведомство', 'СК, ДЮСШ, СДЮСШОР',
            'Ф.И.О. тренера', 'Занятое место'
                ]            
        cells = {}
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=20):
            for cell in row:
                if cell.value in values:
                    cells[cell.value] = cell.coordinate
        
        data = {}
        doyang = ''
        for b in range(ws.min_row + 1, ws.max_row + 1):
            
            fio = ws.cell(b, ws[cells['Ф.И.О.']].column).value
            coach = ws.cell(b, ws[cells['Ф.И.О. тренера']].column).value
            stroka_kat = ws.cell(b, ws[cells['№']].column).value
            SK = ws.cell(b, ws[cells['СК, ДЮСШ, СДЮСШОР']].column).value
            print(stroka_kat)
            if fio is None:
                try:
                    if stroka_kat and len(stroka_kat) < 12:
                        if re.search(r'\d', stroka_kat):
                            doyang = f'{re.search(r'\d', stroka_kat)[0]} доянг'
                        else:
                            doyang = stroka_kat
                        data[doyang] = {}
                    else:
                        kategory = stroka_kat
                        if kategory:
                            data[doyang][kategory] = {}
                except:
                    continue
            if fio and fio !='Ф.И.О.' and doyang != '':
                data[doyang][kategory][fio] = [coach, SK]
        directory_name = 'Круговые'
        try:
            os.mkdir(directory_name)
        except:
            import shutil
            shutil.rmtree(directory_name)
            os.mkdir(directory_name)

        for doyang, kategorii in data.items():
            if f'Круговые/{doyang}' in os.listdir():
                pass
            else:
                os.mkdir(f'Круговые/{doyang}')
            for kategoria, sportsman in kategorii.items():
                print(sportsman.items())
                wb_kat = openpyxl.load_workbook('Круговая.xlsx')
                print(wb_kat.sheetnames)
                if len(sportsman.items()) < 4:
                    ws_kat = wb_kat['3']
                elif len(sportsman.items()) == 4:
                    ws_kat = wb_kat['4']
                else:
                    ws_kat = wb_kat['5']
                for i, (fio, (coach, SK)) in enumerate(sportsman.items(), start=2):
                    ws_kat.cell(i, 1).value = fio
                    ws_kat.cell(i, 2).value = f'{coach} | {SK}'
                wb_kat.save(f'Круговые/{doyang}/{kategoria}.xlsx')

create_circular_kategory(file)
                    
                    



                
                
